#coding: utf-8
#! python2
#HoursTrackerStat (Main)
#WeekTracker week
import csv
import appex
import codecs
import calendar
from datetime import timedelta, date
from math import ceil
import datetime
import console
#############################################
#   EXCEL BEGIN
#############################################
import openpyxl
from openpyxl.styles import Font, Side, Border, Alignment, PatternFill

import holidays
feriados = holidays.Portugal()

class TipoCelula:
    Cima, Baixo = range(2)
fill_ferias = PatternFill(fill_type='solid', start_color='7F7F7F', end_color='7F7F7F')

fill_feriado = PatternFill(fill_type='solid', start_color='3A81FF', end_color='3A81FF')

fill_fds = PatternFill(fill_type='solid', start_color='bcbcbc', end_color='bcbcbc')
#
thick = Side(border_style="thick", color ="000000")
thin = Side(border_style="thin", color ="000000")

border_thick_top = Border(top = thick, bottom= thin, left=thin, right=thin)
border_thick_bottom = Border(top = thin, bottom= thick, left=thin, right=thin)

border_thick_top_left = Border(top = thick, bottom= thin, left=thick, right=thin)
border_thick_top_right = Border(top = thick, bottom= thin, left=thin, right=thick)

border_mes = Border(top = thick, bottom= thick, left=thick, right=thick)

header_pos={}
header_pos['b']=0
header_pos['c']=1
header_pos['d']=2
header_pos['e']=3
header_pos['f']=4
header_pos['g']=5
header_pos['h']=6

header_letter={}
header_letter['0']='b'
header_letter['1']='c'
header_letter['2']='d'
header_letter['3']='e'
header_letter['4']='f'
header_letter['5']='g'
header_letter['6']='h'

header_pos_inicio_semana=0 # border_thick_left
header_pos_fim_semana=6 # border_thick_right

tabela_ultima_linha =14;

'''class CelulaTipo(enum):
    cima = 0
    baixo = 1
    '''
def header(celula_letra, celula_num, texto, border, fill = None):
  global sheet
  
  c = celula_letra+str(celula_num)
  sheet[c].value = texto
  sheet[c].border = border
  if fill is not None:
    sheet[c].fill = fill
    
def celula(tipoCelula, celula_letra, celula_num, texto, fill=None):
  global sheet
    
  c = celula_letra+str(celula_num)
  sheet[c].value = texto
  cborder = Border()  
  align = Alignment()
  #todo centrar texto  
  if tipoCelula == TipoCelula.Cima:
    align.horizontal = "right"
    if celula_num ==5:
      cborder.top =thick
    else:
      cborder.top =thin
  elif tipoCelula == TipoCelula.Baixo:
    align.horizontal = "center"
    if celula_num == tabela_ultima_linha:
      cborder.bottom = thick
    else:
      cborder.bottom = thin
      
  if header_pos[celula_letra] == header_pos_inicio_semana:
    cborder.left=thick
    cborder.right=thin
  elif header_pos[celula_letra] == header_pos_fim_semana:
    cborder.right=thick
    cborder.left=thin
  elif header_pos[celula_letra] == header_pos_fim_semana-1:
    cborder.right=thin
    cborder.left=thick
  else:
    cborder.right=thin
    cborder.left=thin
    
  sheet[c].border = cborder
  if fill is not None:
    sheet[c].fill = fill
  sheet[c].alignment = align
  
def celula_vazia(celula_letra, celula_num):
  celula(TipoCelula.Cima, celula_letra, celula_num, '')
  celula(TipoCelula.Baixo, celula_letra, celula_num+1, '')  

def celula_ok(celula_letra, celula_num, texto):
  celula(TipoCelula.Cima, celula_letra, celula_num, texto)
  celula(TipoCelula.Baixo, celula_letra, celula_num+1, 'ok')  

def celula_ferias(celula_letra, celula_num, texto):
  celula(TipoCelula.Cima, celula_letra, celula_num, texto, fill_ferias)
  celula(TipoCelula.Baixo, celula_letra, celula_num+1, 'Ferias', fill_ferias)  
  
def celula_feriado(celula_letra, celula_num, texto):
  celula(TipoCelula.Cima, celula_letra, celula_num, texto, fill_feriado)
  celula(TipoCelula.Baixo, celula_letra, celula_num+1, 'Feriado', fill_feriado)    
  
def celula_fds(celula_letra, celula_num, texto):
  celula(TipoCelula.Cima, celula_letra, celula_num, texto, fill_fds)
  celula(TipoCelula.Baixo, celula_letra, celula_num+1, '',fill_fds)   

def header_excel():
  #Header, Segunda, terca, etc
  header('b',4,'Segunda',border_thick_top_left )
  header('c',4,'Terça' , border_thick_top)
  header('d',4,'Quarta', border_thick_top )
  header('e',4,'Quinta', border_thick_top )
  header('f',4,'Sexta', border_thick_top )
  header('g',4,'Sabado', Border(top = thick, bottom= thin, left=thick, right=thin), fill_fds )
  header('h',4,'Domingo',border_thick_top_right, fill_fds )
  
  
def test_excel():
  celula_vazia('b',5)
  celula_ok('c',5,'1')
  celula_ok('d',5,'2')
  celula_ok('e',5,'3')
  celula_ok('f',5,'4')
  celula_fds('g', 5, '5')
  celula_fds('h', 5, '6')

  celula_ok('b',7,'7')
  celula_ok('c',7,'8')
  celula_ok('d',7,'9')
  celula_ok('e',7,'10')
  celula_ok('f',7,'11')
  celula_fds('g', 7, '12')
  celula_fds('h', 7, '13')

  celula_ok('b',9,'14')
  celula_feriado('c',9,'15')
  celula_ok('d',9,'16')
  celula_ok('e',9,'17')
  celula_ok('f',9,'18')
  celula_fds('g', 9, '19')
  celula_fds('h', 9, '20')

  celula_ferias('b',11,'21')
  celula_ferias('c',11,'22')
  celula_ferias('d',11,'23')
  celula_ferias('e',11,'24')
  celula_ferias('f',11,'25')
  celula_fds('g',11, '26')
  celula_fds('h',11, '27')


  celula_ferias('b',13,'28')
  celula_ferias('c',13,'29')
  celula_ferias('d',13,'30')
  celula_ferias('e',13,'31')
  celula_vazia('f',13)
  celula_fds('g',13,'')
  celula_fds('h',13,'')
#############################################
#   EXCEL END
#############################################  

def get_month_range(start_date = None):
    if start_date is None:
        start_date = date.today()
    print ('startdate:{} '.format(start_date))
    start_date = start_date.replace(day = 1)        
    print ('startdate:{} '.format(start_date))
    days_in_month = calendar.monthrange(start_date.year, start_date.month)[1]
    
    end_date = start_date + timedelta(days=days_in_month)
    return  (start_date,end_date )

csv.register_dialect('HoursTtracker-csv', delimiter=',', quoting=csv.QUOTE_ALL)
from WeekTracker import WeekTracker
from MonthTracker import MonthTracker

file_to_open='CSVExport_setembro.csv'
        
#for ios
if appex.is_running_extension():
  file_paths = appex.get_file_paths()
  for i, file in enumerate(file_paths):
    if file.endswith('/CSVExport.csv'):
    	print(file)
    	file_to_open=file
csv_file = codecs.open(file_to_open,'r','utf-8')
#for windows
#csv_file = open(file_to_open)
#print(csv_file.read())       
print(file_to_open)
csv_reader = csv.reader(csv_file,'HoursTtracker-csv')

month = None
for line, row in enumerate(csv_reader):
    '''
        "Job","Clocked In","Clocked Out","Duration","Comment","Tags","Breaks","Adjustments","TotalTimeAdjustment","TotalEarningsAdjustment"
        "SIBS","21/02/17 12:51","21/02/17 17:12","4,35","Mastercard pin offline autorizacoes  Recibo Bits","ARCTIC - Construction - Suporte interno/externo sem SW","","","",""
    '''
    if line >1:
        print("line {} ".format(line))
        job = row[0]
        clockedIn = row[1]
        clockedOut = row[2]
        duration = str(row[3])
        comments = row[4]     
        tags     = row[5]
        print ('job:{} clockedIn:{}, clockedOut:{} , duration:{} , comments:{} , tags:{}clockedIn '.format(job, clockedIn, clockedOut, duration, comments, tags))
        if month is None:        
            month = MonthTracker(clockedIn)
        month.add(job, clockedIn, clockedOut, duration, comments, tags)
    else:
        print(row)

if month.numWeeks() == 6:
  tabela_ultima_linha = 16
elif month.numWeeks() == 5:
  tabela_ultima_linha = 14
else:
  tabela_ultima_linha = 12
 
print('tabela_ultima_linha',tabela_ultima_linha)
wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.get_active_sheet()



sheet.title = 'Calendario de presencas'
sn =wb.get_sheet_names()
print(sn)
sheet = wb.get_active_sheet()

header_excel()        
        
a_day = timedelta(days=1)
print(month.dt)
first_day, last_day = get_month_range(month.dt)
print('first_day:{} last_day:{}'.format(first_day,last_day))
month.calcHours()
dayNameList= ['Segunda', 'Terça', 'Quarta','Quinta','Sexta','Sábado','Domingo']
weekday = 0
row = 5

today = datetime.datetime.now()
first_weekday = first_day.weekday()
first_day_letter = header_letter[str(first_weekday)]
while weekday < first_weekday:
  print('first_weekday:{} first_day_letter:{} current weekday:{}'.format(first_weekday, first_day_letter, weekday))
  celula_vazia(header_letter[str(weekday)],row)
  weekday+=1

while first_day < last_day:         
  first_weekday = first_day.weekday()  
  if first_weekday == 0:
    print('weekday:{} first_weekday:{}. Incrementing ROW'.format(weekday, first_weekday))    
    row+=2
  print('first_weekday:{} row:{} letter:{}'.format(first_weekday,row, header_letter[str(first_weekday)]))
  print("day:{} DayName:{} dayTotalTimeRounded:{}".format(first_day,dayNameList[first_day.weekday()], month.getHoursOfDay(first_day)))
  month_day = month.getDayTracker(first_day)
  if month_day is not None:    
    if month_day.dayTotalTimeRounded > 0:
      celula_ok(header_letter[str(first_weekday)],row,str(first_day.day))
    elif date(first_day.year,first_day.month, first_day.day) in feriados:
      print('feriado detectado')
      celula_feriado(header_letter[str(first_weekday)],row,str(first_day.day))
    elif datetime.datetime(first_day.year,first_day.month, first_day.day) >= today:
      celula_ok(header_letter[str(first_weekday)],row,str(first_day.day))
    else:
      print('ferias else')
      celula_ferias(header_letter[str(first_weekday)],row,str(first_day.day))
  elif first_weekday in [5,6]:
      celula_fds(header_letter[str(first_weekday)], row, str(first_day.day))
  first_day += a_day

#write empty cells until end of week
last_weekday = last_day.weekday()
while last_weekday < 7:
  if last_weekday in [5,6]:
    celula_fds(header_letter[str(last_weekday)], row, '')
  else:
    celula_vazia(header_letter[str(last_weekday)],row)
  last_weekday+=1
mes={}
mes['1']='Janeiro'
mes['2']='FEvereiro'
mes['3']='Março'
mes['4']='Abril'
mes['5']='Maio'
mes['6']='Junho'
mes['7']='Julho'
mes['8']='Agosto'
mes['9']='Setembro'
mes['10']='Outubro'
mes['11']='Novembro'
mes['12']='Dezembro'

fonte_def = Font(name = 'calibri', size= 16, bold= True)
sheet['b2'].value =mes[str(month.dt.month)]+' 2017'
sheet['b2'].font = fonte_def

excel_file = 'David Jorge - {}-{:02} {}.xlsx'.format(month.dt.year,month.dt.month, mes[str(month.dt.month)])
print('excel_file:',excel_file)
wb.save(excel_file)
console.open_in(excel_file)
